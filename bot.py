"""
Homework / Conspect Bot (stable single-file version)

Features:
- SQLite storage (users, submissions, miss_reasons)
- Submit homework or conspect (text / photo / album)
- Save conspects to disk under CONSPECTS_DIR/<user_id>/
- Daily: at 23:55 admin receives Excel with all users' daily statuses (cumulative rows)
- Daily: at 23:57 bot asks users who didn't submit that day the reason and stores it
- Admin panel: export user, delete user (by id or username), send today's report immediately, reset all data
- All handlers and DB calls protected; no nested handlers; proper async/await usage
"""
import random
import os
import io
import time
import zipfile
import logging
import shutil
import sqlite3
import unicodedata
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv

load_dotenv()

import pandas as pd

# Additional imports for charts and PNG generation
import matplotlib

matplotlib.use('Agg')  # non-interactive backend for servers
import matplotlib.pyplot as plt
from io import BytesIO as _BytesIO
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from apscheduler.schedulers.asyncio import AsyncIOScheduler

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("homework-bot")

# additional imports for Excel styling and reports
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Reports directory
REPORTS_DIR = os.environ.get('REPORTS_DIR', 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)

# ---------------- CONFIG ----------------
API_TOKEN = os.environ.get("API_TOKEN")
if not API_TOKEN:
    raise ValueError("API_TOKEN is missing. Set it in environment variables!")

# Исправляем получение ADMIN_ID
ADMIN_ID_STR = os.environ.get("ADMIN_ID", "").strip()
if ADMIN_ID_STR and ADMIN_ID_STR.isdigit():
    ADMIN_ID = int(ADMIN_ID_STR)
else:
    ADMIN_ID = 0
    logger.warning("ADMIN_ID is not set or invalid. Set ADMIN_ID env var (your Telegram id).")

DB_FILE = os.environ.get("DB_FILE", "bot.db")
CONSPECTS_DIR = os.environ.get("CONSPECTS_DIR", "conspects")
os.makedirs(CONSPECTS_DIR, exist_ok=True)

# Basic validation
if not API_TOKEN:
    logger.warning("API_TOKEN is empty. Set API_TOKEN env var before running.")
if ADMIN_ID == 0:
    logger.warning("ADMIN_ID is 0 or not set. Set ADMIN_ID env var (your Telegram id).")

# ---------------- BOT / DISPATCHER / SCHEDULER ----------------
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
scheduler = AsyncIOScheduler()

# in-memory flows
pending: Dict[str, Dict[str, Any]] = {}  
admin_pending: Dict[str, Dict[str, Any]] = {} 
media_groups: Dict[str, Dict[str, Any]] = {}  
reasons_pending: Dict[str, str] = {}  


# ---------------- UTIL ----------------
def today_str() -> str:
    return date.today().isoformat()


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def slugify_filename(s: str) -> str:
    s = unicodedata.normalize('NFKD', s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("/", "_").replace("\\", "_")
    allowed = "-_.() abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 "
    return "".join(c for c in s if c in allowed)[:200] or "file"


def parse_date(text: str) -> Optional[date]:
    text = (text or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    return None


# ---------------- USERNAME HELPERS ----------------
def get_username_by_id(user_id: int) -> str:
    
    try:
        cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
        result = cursor.fetchone()
        return result[0] if result else ""
    except Exception as e:
        logger.exception(f"Error getting username for user_id {user_id}")
        return ""


def get_user_display_name_by_id(user_id: int) -> str:
    
    try:
        cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (user_id,))
        result = cursor.fetchone()
        if result:
            username, first_name = result
            return username if username else (first_name if first_name else f"user_{user_id}")
        return f"user_{user_id}"
    except Exception as e:
        logger.exception(f"Error getting display name for user_id {user_id}")
        return f"user_{user_id}"


def mention_html(user: types.User = None, user_id: int = None) -> str:
    
    if user:
        name = (user.username and f"@{user.username}") or (user.first_name or f"user_{user.id}")
        return f"<a href='tg://user?id={user.id}'>{name}</a>"
    elif user_id:
        display_name = get_user_display_name_by_id(user_id)
        return f"<a href='tg://user?id={user_id}'>{display_name}</a>"
    else:
        return "Неизвестный пользователь"


def mention_html_by_id(user_id: int) -> str:
    
    display_name = get_user_display_name_by_id(user_id)
    return f"<a href='tg://user?id={user_id}'>{display_name}</a>"


# ---------------- DB ----------------
conn = sqlite3.connect(DB_FILE, check_same_thread=False)
cursor = conn.cursor()


def init_db():
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        username TEXT,
        first_name TEXT
    )
    ''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS submissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        type TEXT,
        section TEXT,
        topic_id TEXT,
        topic_title TEXT,
        content_type TEXT,
        content_summary TEXT,
        photo_file_id TEXT,
        message_id INTEGER,
        date TEXT,
        ts TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    ''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS miss_reasons (
        user_id INTEGER,
        date TEXT,
        reason TEXT,
        PRIMARY KEY (user_id, date)
    )
    ''')
    conn.commit()


init_db()

# ---------------- TOPICS ----------------
SECTIONS = {
    "Основы Питона": [
        {"id": "op1", "title": "Вводный урок", "dz_allowed": True},
        {"id": "op2", "title": "Условия и целочисленные операторы", "dz_allowed": False},
        {"id": "op3", "title": "Цикл for", "dz_allowed": True},
        {"id": "op4", "title": "Цикл while", "dz_allowed": True},
        {"id": "op5", "title": "Практика: цикл for и while", "dz_allowed": False},
        {"id": "op6", "title": "Строки и срезы", "dz_allowed": False},
        {"id": "op7", "title": "Списки и генераторы", "dz_allowed": True},
    ],
    "ЕГЭ 1-27": [
        {"id": f"ege{n}", "title": f"Задание {n}", "dz_allowed": True} for n in range(1, 28)
    ]
}


# USER / SUBMISSION HELPERS
def ensure_user_record_obj(user: types.User):
    try:
        cursor.execute('INSERT OR IGNORE INTO users (id, username, first_name) VALUES (?, ?, ?)',
                       (user.id, user.username or "", user.first_name or ""))
        cursor.execute('UPDATE users SET username = ?, first_name = ? WHERE id = ?',
                       (user.username or "", user.first_name or "", user.id))
        conn.commit()
    except Exception:
        logger.exception("ensure_user_record_obj error")


def ensure_user_record_by_id(uid: int, username: str = "", first_name: str = ""):
    try:
        cursor.execute('INSERT OR IGNORE INTO users (id, username, first_name) VALUES (?, ?, ?)',
                       (uid, username, first_name))
        cursor.execute('UPDATE users SET username = ?, first_name = ? WHERE id = ?',
                       (username or "", first_name or "", uid))
        conn.commit()
    except Exception:
        logger.exception("ensure_user_record_by_id error")


async def ensure_user_with_current_data(user_id: int, bot_instance: Bot = None):
    
    try:
        if bot_instance is None:
            bot_instance = bot

        # Пытаемся получить актуальные данные пользователя из Telegram
        try:
            chat_member = await bot_instance.get_chat_member(user_id, user_id)
            user = chat_member.user
        except:
            # Если не получается через get_chat_member, используем существующие данные из БД
            cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (user_id,))
            result = cursor.fetchone()
            if result:
                username, first_name = result
                return
            else:
                # Если нет данных в БД, создаем базовую запись
                cursor.execute('INSERT OR IGNORE INTO users (id, username, first_name) VALUES (?, ?, ?)',
                               (user_id, "", f"user_{user_id}"))
                conn.commit()
                return

        # Обновляем запись в базе данных
        cursor.execute(
            'INSERT OR REPLACE INTO users (id, username, first_name) VALUES (?, ?, ?)',
            (user.id, user.username or "", user.first_name or "")
        )
        conn.commit()
        logger.info(f"Updated user data: {user.id}, @{user.username}, {user.first_name}")

    except Exception as e:
        logger.exception(f"Error ensuring user data for {user_id}")


def add_submission_obj(user: types.User, submission: dict, message_id: int = None):
    ensure_user_record_obj(user)
    try:
        cursor.execute('''
            INSERT INTO submissions (
                user_id, type, section, topic_id, topic_title, content_type, 
                content_summary, photo_file_id, message_id, date, ts
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            user.id, submission['type'], submission['section'], submission['topic_id'],
            submission['topic_title'], submission['content_type'], submission.get('content_summary', ''),
            submission.get('photo_file_id', ''), message_id, submission['date'], submission['ts']
        ))
        conn.commit()
    except Exception:
        logger.exception("add_submission_obj error")


# FILE SAVE
async def download_file_bytes(file_id: str) -> Optional[bytes]:
    try:
        f = await bot.get_file(file_id)
        file_path = f.file_path
        b = await bot.download_file(file_path)
        if hasattr(b, "read"):
            return b.read()
        return b
    except Exception:
        logger.exception("download error")
        return None


def save_conspect_files(user_id: str, section: str, topic_id: str, files: List[bytes], filenames: List[str]):
    base = os.path.join(CONSPECTS_DIR, str(user_id), f"{slugify_filename(section)}_{slugify_filename(topic_id)}")
    ensure_dir(base)
    saved = []
    for content, name in zip(files, filenames):
        safe = slugify_filename(name)
        path = os.path.join(base, safe)
        try:
            with open(path, "wb") as f:
                f.write(content)
            saved.append(path)
        except Exception:
            logger.exception("Failed to save file %s", path)
    return saved


def save_conspect_text(user_id: str, section: str, topic_id: str, text: str):
    base = os.path.join(CONSPECTS_DIR, str(user_id), f"{slugify_filename(section)}_{slugify_filename(topic_id)}")
    ensure_dir(base)
    fname = datetime.utcnow().strftime("%Y%m%d_%H%M%S") + ".txt"
    path = os.path.join(base, fname)
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
        return path
    except Exception:
        logger.exception("Failed to save text file %s", path)
        return None


# ---------------- KEYBOARDS ----------------
def make_main_kb(is_admin: bool):
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton("📚 Сдать ДЗ"))
    kb.add(KeyboardButton("📘 Сдать конспект"))
    kb.row(KeyboardButton("📁 Мои конспекты"))
    kb.row(KeyboardButton("📌 Главное меню"))
    if is_admin:
        kb.row(KeyboardButton("🛠️ Админ-панель"))
    return kb


def section_keyboard():
    kb = InlineKeyboardMarkup()
    for sec in SECTIONS.keys():
        kb.add(InlineKeyboardButton(sec, callback_data=f"sec|{sec}"))
    return kb


def topics_keyboard(section_name):
    kb = InlineKeyboardMarkup(row_width=2)
    topics = SECTIONS.get(section_name, [])
    for t in topics:
        kb.add(InlineKeyboardButton(t["title"], callback_data=f"topic|{section_name}|{t['id']}"))
    kb.add(InlineKeyboardButton("📊 Исторический отчёт (ALL)", callback_data="admin|full_history_manual"))
    kb.add(InlineKeyboardButton("Отмена", callback_data="cancel"))
    return kb


def admin_kb():
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("📋 Дневной отчёт (подробный)", callback_data="admin|daily_full"))
    kb.add(InlineKeyboardButton("🆕 Создать таблицу (сохранить)", callback_data="admin|new_report"))
    kb.add(InlineKeyboardButton("🗑️ Удалить отчёты", callback_data="admin|delete_reports"))
    kb.add(InlineKeyboardButton("🧹 Очистить пустые столбцы", callback_data="admin|cleanup_columns"))
    kb.add(InlineKeyboardButton("📤 Выслать таблицу сейчас", callback_data="admin|send_daily_now"))
    kb.add(InlineKeyboardButton("👤 Выгрузить ученика (Excel + фото ZIP)", callback_data="admin|export_user"))
    kb.add(InlineKeyboardButton("🗑️ Удалить ученика (полностью)", callback_data="admin|delete_user"))
    kb.add(InlineKeyboardButton("♻️ Сбросить все статусы", callback_data="admin|reset_all"))
    kb.add(InlineKeyboardButton("📊 Исторический отчёт (ALL)", callback_data="admin|full_history_manual"))
    kb.add(InlineKeyboardButton("Отмена", callback_data="cancel"))
    return kb


# PRAISE
generic_praise = ["Молодец, отличная работа!", "Здорово, так держать!", "Круто, ты справился!", "Умница, ДЗ принято!","АЙ ЛЕВ","Лёва оценил!!!","Ты - будущий 100-балльник"]
context_praise_templates = ["Отлично поработал над «{topic}» — заметен прогресс!",
                            "Круто! Тема «{topic}» покоряется тебе всё лучше.",
                            "Хорошая работа по «{topic}» — продолжай в том же духе!"]
photo_praise = ["Фото принято — выглядит аккуратно!", "Классный снимок, спасибо!",
                "Фото получено — спасибо за старание!", "Имба, Леве понравится!" , "Ну ты прям машина!!"]


def get_praise(user_id: int, section: str, topic_title: str, content_type: str) -> str:
    try:
        cursor.execute('SELECT COUNT(*) FROM submissions WHERE user_id = ?', (user_id,))
        total_subs = cursor.fetchone()[0]
    except Exception:
        total_subs = 0
    messages = []
    if topic_title:
        messages.append(random.choice(context_praise_templates).format(topic=topic_title))
    if content_type == "photo":
        messages.append(random.choice(photo_praise))
    else:
        messages.append(random.choice(generic_praise))
    if total_subs >= 10:
        messages.append("Ты постоянный участник — это впечатляет! 🔥")
    elif total_subs >= 3:
        messages.append("Отлично, ты активно сдаёшь. Продолжай!")
    n = 2 if random.random() > 0.4 else 1
    return " ".join(messages[:n])


# MAKING EXCEL
def style_worksheet(ws):
    header_fill = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header row styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # apply border to body and compute column widths
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # auto width
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 5, 60)


def submissions_for_date(target_date: date) -> List[dict]:
    """Build summary rows including a task_flag (topic_id) and a mention link fallback."""
    dstr = target_date.isoformat()
    try:
        cursor.execute('SELECT id, username, first_name FROM users ORDER BY id')
        users = cursor.fetchall()
        result = []
        for u in users:
            uid, uname, fname = u
            display_name = uname or fname or ""
            # compute types submitted
            cursor.execute('SELECT type FROM submissions WHERE user_id = ? AND date = ?', (uid, dstr))
            types = {r[0] for r in cursor.fetchall()}
            dz = 'dz' in types
            cons = 'conspect' in types
            # miss reason
            cursor.execute('SELECT reason FROM miss_reasons WHERE user_id = ? AND date = ?', (uid, dstr))
            rr = cursor.fetchone()
            reason = rr[0] if rr else ""
            # last submission topic (task flag) for that day
            cursor.execute(
                'SELECT topic_id, topic_title, message_id FROM submissions WHERE user_id = ? AND date = ? ORDER BY ts DESC LIMIT 1',
                (uid, dstr))
            trow = cursor.fetchone()
            task_flag = trow[0] if trow and trow[0] else (trow[1] if trow and trow[1] else "")
            message_id = trow[2] if trow and len(trow) > 2 and trow[2] else None
            result.append({
                "user_id": uid,
                "username": display_name,
                "date": dstr,
                "dz_submitted": int(dz),
                "conspect_submitted": int(cons),
                "miss_reason": reason,
                "task_flag": task_flag,
                "message_id": message_id
            })
        return result
    except Exception:
        logger.exception("submissions_for_date build error")
        return []


def make_daily_excel(target_date: date) -> io.BytesIO:
    rows = submissions_for_date(target_date)
    df_summary = pd.DataFrame(rows)
    try:
        cursor.execute(
            "SELECT s.user_id, u.username, u.first_name, s.type, s.section, s.topic_id, s.topic_title, s.content_type, s.content_summary, s.photo_file_id, s.date, s.ts "
            "FROM submissions s JOIN users u ON s.user_id = u.id "
            "WHERE s.date = ? "
            "ORDER BY s.ts",
            (target_date.isoformat(),)
        )
        raw = cursor.fetchall()
        raw_rows = [{
            "user_id": r[0],
            "username": r[1] or r[2] or "",
            "type": r[3],
            "section": r[4],
            "topic_id": r[5],
            "topic_title": r[6],
            "content_type": r[7],
            "content_summary": r[8],
            "photo_file_id": r[9],
            "date": r[10],
            "ts": r[11]
        } for r in raw]
        df_raw = pd.DataFrame(raw_rows)
    except Exception:
        logger.exception("make_daily_excel raw fetch error")
        df_raw = pd.DataFrame(
            columns=["user_id", "username", "type", "section", "topic_id", "topic_title", "content_type",
                     "content_summary", "photo_file_id", "date", "ts"])

    # ensure task_flag exists in summary (comes from submissions_for_date)
    if "task_flag" not in df_summary.columns:
        df_summary["task_flag"] = ""

    # ensure task_flag in raw (from topic_id)
    if "task_flag" not in df_raw.columns:
        df_raw["task_flag"] = df_raw.get("topic_id", "")

    # write initial excel to BytesIO
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="daily_summary", index=False)
        df_raw.to_excel(writer, sheet_name="raw_submissions", index=False)
    bio.seek(0)

    # post-process with openpyxl to create hyperlinks and style
    wb = load_workbook(filename=io.BytesIO(bio.read()))

    # helper to make display name
    def _display_name(username, first_name, uid):
        if username and str(username).strip():
            return f"@{username}"
        if first_name and str(first_name).strip():
            return first_name
        return f"user_{uid}"

    # process daily_summary sheet
    if "daily_summary" in wb.sheetnames:
        ws = wb["daily_summary"]
        # map headers
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        uid_col = headers.get("user_id")
        uname_col = headers.get("username")
        task_col = headers.get("task_flag")
        if uid_col and uname_col:
            for row in range(2, ws.max_row + 1):
                uid_cell = ws.cell(row=row, column=uid_col)
                uname_cell = ws.cell(row=row, column=uname_col)
                try:
                    uid_val = int(uid_cell.value)
                except Exception:
                    uid_val = None
                if uid_val:
                    try:
                        cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (int(uid_val),))
                        r = cursor.fetchone()
                        uname_db = r[0] if r else None
                        fname_db = r[1] if r else None
                    except Exception:
                        uname_db = None
                        fname_db = None
                    final_display = _display_name(uname_db, fname_db, uid_val)
                    uname_cell.value = final_display
                    uname_cell.hyperlink = f"tg://user?id={uid_val}"
                    uname_cell.style = "Hyperlink"
        style_worksheet(ws)

    # process raw_submissions sheet
    if "raw_submissions" in wb.sheetnames:
        ws2 = wb["raw_submissions"]
        headers2 = {cell.value: idx + 1 for idx, cell in enumerate(ws2[1])}
        uid_col2 = headers2.get("user_id")
        uname_col2 = headers2.get("username")
        topic_id_col = headers2.get("topic_id")
        task_col2 = headers2.get("task_flag")
        if uid_col2 and uname_col2:
            for row in range(2, ws2.max_row + 1):
                uid_cell = ws2.cell(row=row, column=uid_col2)
                uname_cell = ws2.cell(row=row, column=uname_col2)
                try:
                    uid_val = int(uid_cell.value)
                except Exception:
                    uid_val = None
                if uid_val:
                    try:
                        cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (int(uid_val),))
                        r = cursor.fetchone()
                        uname_db = r[0] if r else None
                        fname_db = r[1] if r else None
                    except Exception:
                        uname_db = None
                        fname_db = None
                    final_display = _display_name(uname_db, fname_db, uid_val)
                    uname_cell.value = final_display
                    uname_cell.hyperlink = f"tg://user?id={uid_val}"
                    uname_cell.style = "Hyperlink"
        # ensure task_flag column exists and filled from topic_id
        if topic_id_col:
            if not task_col2:
                task_col_idx = ws2.max_column + 1
                ws2.cell(row=1, column=task_col_idx, value="task_flag")
            else:
                task_col_idx = task_col2
            for row in range(2, ws2.max_row + 1):
                topic_cell = ws2.cell(row=row, column=topic_id_col)
                val = topic_cell.value or ""
                ws2.cell(row=row, column=task_col_idx, value=val)
        style_worksheet(ws2)

    # save workbook to BytesIO and also save a copy on disk
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    # save on disk
    fname = f"daily_report_{target_date.isoformat()}_{int(time.time())}.xlsx"
    fpath = os.path.join(REPORTS_DIR, fname)
    with open(fpath, "wb") as f:
        f.write(out.getbuffer())
    logger.info(f"Saved report to {fpath}")

    out.seek(0)
    return out


async def send_daily_excel_to_admin(target_date: date):
    try:
        bio = make_daily_excel(target_date)
        fname = f"daily_report_{target_date.isoformat()}.xlsx"
        await bot.send_document(ADMIN_ID, InputFile(bio, filename=fname))
        await bot.send_message(ADMIN_ID, f"Отчёт за {target_date.isoformat()} отправлен.")
    except Exception:
        logger.exception("send_daily_excel_to_admin error")
        try:
            await bot.send_message(ADMIN_ID, "Ошибка при формировании отчёта.")
        except Exception:
            pass


# HANDLERS
@dp.message_handler(commands=["start", "menu"])
async def cmd_start(message: types.Message):
    ensure_user_record_obj(message.from_user)
    is_admin = (message.from_user.id == ADMIN_ID)
    await message.answer("Привет! Я бот для сдачи ДЗ и конспектов.\nВыбери действие:",
                         reply_markup=make_main_kb(is_admin))


@dp.message_handler(lambda m: m.text == "📚 Сдать ДЗ")
async def cmd_send_dz(message: types.Message):
    pending[str(message.from_user.id)] = {"type": "dz", "ts": datetime.utcnow().isoformat()}
    await message.answer("Выбери раздел:", reply_markup=section_keyboard())


@dp.message_handler(lambda m: m.text == "📘 Сдать конспект")
async def cmd_send_conspect(message: types.Message):
    pending[str(message.from_user.id)] = {"type": "conspect", "ts": datetime.utcnow().isoformat()}
    await message.answer("Выбери раздел:", reply_markup=section_keyboard())


@dp.message_handler(lambda m: m.text == "📁 Мои конспекты")
async def cmd_my_conspects(message: types.Message):
    uid = str(message.from_user.id)
    folder = os.path.join(CONSPECTS_DIR, uid)
    if not os.path.exists(folder):
        return await message.answer("У тебя пока нет сохранённых конспектов.",
                                    reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder):
            for f in files:
                full = os.path.join(root, f)
                arc = os.path.relpath(full, folder)
                zf.write(full, arc)
    buf.seek(0)
    await bot.send_document(message.chat.id, InputFile(buf, filename="my_conspects.zip"),
                            reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))


@dp.message_handler(lambda m: m.text == "📌 Главное меню")
async def cmd_main_menu(message: types.Message):
    is_admin = (message.from_user.id == ADMIN_ID)
    await message.answer("Главное меню:", reply_markup=make_main_kb(is_admin))


@dp.message_handler(lambda m: m.text == "🛠️ Админ-панель")
async def cmd_admin_panel(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return await message.answer("Доступ запрещён.", reply_markup=make_main_kb(False))
    # show admin inline keyboard
    await message.answer("Админ-панель:", reply_markup=admin_kb())


# Section / topic callbacks
@dp.callback_query_handler(lambda c: c.data and c.data.startswith("sec|"))
async def cb_section_choose(call: types.CallbackQuery):
    _, section = call.data.split("|", 1)
    uid = str(call.from_user.id)
    if uid not in pending:
        await call.answer("Нет активного действия. Нажми 'Сдать ДЗ' или 'Сдать конспект' в меню.", show_alert=True)
        return
    pending[uid]["section"] = section
    await bot.send_message(call.from_user.id, f"Выбран раздел: {section}\nВыбери тему:",
                           reply_markup=topics_keyboard(section))
    await call.answer()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith("topic|"))
async def cb_topic_choose(call: types.CallbackQuery):
    _, section, topic_id = call.data.split("|", 2)
    uid = str(call.from_user.id)
    if uid not in pending or "section" not in pending[uid]:
        await call.answer("Нет активного действия. Сначала нажми 'Сдать ДЗ' или 'Сдать конспект'.", show_alert=True)
        return
    topic = None
    for t in SECTIONS.get(section, []):
        if t["id"] == topic_id:
            topic = t
            break
    if not topic:
        await call.answer("Тема не найдена.", show_alert=True)
        return
    if pending[uid]["type"] == "dz" and not topic.get("dz_allowed", True):
        await call.answer("Для этой темы ДЗ не предусмотрено. Выбери другую тему или сдай конспект.", show_alert=True)
        return
    pending[uid]["topic"] = {"id": topic["id"], "title": topic["title"]}
    await bot.send_message(call.from_user.id,
                           f"Тема: {topic['title']}\nТеперь отправь {('ДЗ' if pending[uid]['type'] == 'dz' else 'конспект')} текстом или фото. В подписи к фото можно указать подробности.",
                           reply_markup=None)
    await call.answer()


@dp.callback_query_handler(lambda c: c.data == "cancel")
async def cb_cancel(call: types.CallbackQuery):
    uid = str(call.from_user.id)
    pending.pop(uid, None)
    admin_pending.pop(uid, None)
    await bot.send_message(call.from_user.id, "Операция отменена.",
                           reply_markup=make_main_kb(call.from_user.id == ADMIN_ID))
    await call.answer()


@dp.message_handler(content_types=types.ContentType.TEXT)
async def handle_text(message: types.Message):
    uid = str(message.from_user.id)
    text = message.text.strip()

    if message.text in ("📊 Статистика",):
        return

    if message.text in ("📚 Сдать ДЗ", "📘 Сдать конспект", "📁 Мои конспекты", "📌 Главное меню"):
        pending.pop(uid, None)

    if uid in reasons_pending:
        miss_date = reasons_pending.pop(uid)
        try:
            cursor.execute('INSERT OR REPLACE INTO miss_reasons (user_id, date, reason) VALUES (?, ?, ?)',
                           (int(uid), miss_date, text))
            conn.commit()
        except Exception:
            logger.exception("saving miss reason failed")
        await message.answer(f"Спасибо — причина пропуска за {miss_date} сохранена.",
                             reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
        return

    # admin flows waiting for text
    if str(message.from_user.id) in admin_pending:
        # handled by admin_pending_text callback
        return

    if uid in pending and "section" in pending[uid] and "topic" in pending[uid]:
        p = pending.pop(uid)
        sub = {"type": p["type"], "section": p["section"], "topic_id": p["topic"]["id"],
               "topic_title": p["topic"]["title"],
               "content_type": "text", "content_summary": (text if len(text) <= 300 else text[:297] + "..."),
               "date": today_str(), "ts": datetime.utcnow().isoformat()}
        add_submission_obj(message.from_user, sub, message.message_id)
        if sub["type"] == "conspect":
            try:
                save_conspect_text(uid, sub["section"].replace("/", "_"), sub["topic_id"], text)
            except Exception:
                logger.exception("Failed to save conspect text")
        praise = get_praise(message.from_user.id, sub["section"], sub["topic_title"], "text")
        await message.answer(praise, reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
        try:
            # Обновляем данные пользователя перед отправкой уведомления админу
            await ensure_user_with_current_data(message.from_user.id)
            await bot.send_message(ADMIN_ID,
                                   f"✅ {mention_html_by_id(message.from_user.id)} прислал {sub['type'].upper()}: {sub['section']} — {sub['topic_title']}\n{sub['content_summary']}",
                                   parse_mode="HTML")
        except Exception:
            pass
        return

    low = text.lower()
    if low.startswith("дз") or low.startswith("конспект"):
        kind = "dz" if low.startswith("дз") else "conspect"
        sub = {"type": kind, "section": "Без раздела", "topic_id": "none", "topic_title": text.split("\n", 1)[0][:50],
               "content_type": "text", "content_summary": (text if len(text) <= 300 else text[:297] + "..."),
               "date": today_str(), "ts": datetime.utcnow().isoformat()}
        add_submission_obj(message.from_user, sub, message.message_id)
        if kind == "conspect":
            try:
                save_conspect_text(uid, "Без_раздела", "none", text)
            except Exception:
                logger.exception("Failed to save conspect text")
        praise = get_praise(message.from_user.id, sub["section"], sub["topic_title"], "text")
        await message.answer("Записал без выбора темы. В следующий раз выбери тему через меню.\n\n" + praise,
                             reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
        try:
            # Обновляем данные пользователя перед отправкой уведомления админу
            await ensure_user_with_current_data(message.from_user.id)
            await bot.send_message(ADMIN_ID,
                                   f"✅ {mention_html_by_id(message.from_user.id)} прислал {sub['type'].upper()} (без темы): {sub['content_summary']}",
                                   parse_mode="HTML")
        except Exception:
            pass
        return

    await message.answer("Чтобы сдать ДЗ или конспект: нажми соответствующую кнопку в меню и выбери тему.",
                         reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))


@dp.message_handler(content_types=types.ContentType.PHOTO)
async def handle_photo(message: types.Message):
    uid = str(message.from_user.id)
    caption = message.caption or ""
    mgid = getattr(message, "media_group_id", None)

    if mgid:
        key = f"{uid}|{mgid}"
        entry = media_groups.get(key)
        file_id = message.photo[-1].file_id
        if not entry:
            media_groups[key] = {"file_ids": [file_id], "caption": caption or "", "last_update": time.time(),
                                 "pending_snapshot": pending.get(uid), "uid": uid}
        else:
            entry["file_ids"].append(file_id)
            if caption:
                entry["caption"] = caption
            entry["last_update"] = time.time()
        await message.answer("Фото получено — обрабатываю альбом...",
                             reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
        return

    # single photo
    if uid in pending and "section" in pending[uid] and "topic" in pending[uid]:
        p = pending.pop(uid)
        file_id = message.photo[-1].file_id
        summary = (caption[:200] + "...") if len(caption) > 200 else caption or "Фото"
        sub = {"type": p["type"], "section": p["section"], "topic_id": p["topic"]["id"],
               "topic_title": p["topic"]["title"],
               "content_type": "photo", "content_summary": summary, "photo_file_id": file_id, "date": today_str(),
               "ts": datetime.utcnow().isoformat()}
        add_submission_obj(message.from_user, sub, message.message_id)
        if sub["type"] == "conspect":
            try:
                fbytes = await download_file_bytes(file_id)
                if fbytes:
                    save_conspect_files(uid, sub["section"].replace("/", "_"), sub["topic_id"], [fbytes],
                                        [f"photo_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.jpg"])
            except Exception:
                logger.exception("Failed to save conspect photo")
        praise = get_praise(message.from_user.id, sub["section"], sub["topic_title"], "photo")
        await message.answer(praise, reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
        try:
            # Обновляем данные пользователя перед отправкой уведомления админу
            await ensure_user_with_current_data(message.from_user.id)
            await bot.send_message(ADMIN_ID,
                                   f"📸 {mention_html_by_id(message.from_user.id)} прислал {sub['type'].upper()}: {sub['section']} — {sub['topic_title']} — {summary}",
                                   parse_mode="HTML")
            await bot.forward_message(ADMIN_ID, message.chat.id, message.message_id)
        except Exception:
            pass
        return

    # caption-start fallback
    if caption.lower().startswith("дз") or caption.lower().startswith("конспект"):
        kind = "dz" if caption.lower().startswith("дз") else "conspect"
        file_id = message.photo[-1].file_id
        summary = (caption[:200] + "...") if len(caption) > 200 else caption
        sub = {"type": kind, "section": "Без раздела", "topic_id": "none",
               "topic_title": summary.split("\n", 1)[0][:50],
               "content_type": "photo", "content_summary": summary, "photo_file_id": file_id, "date": today_str(),
               "ts": datetime.utcnow().isoformat()}
        add_submission_obj(message.from_user, sub, message.message_id)
        if kind == "conspect":
            try:
                fbytes = await download_file_bytes(file_id)
                if fbytes:
                    save_conspect_files(uid, "Без_раздела", "none", [fbytes],
                                        [f"photo_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.jpg"])
            except Exception:
                logger.exception("Failed to save conspect photo")
        praise = get_praise(message.from_user.id, sub["section"], sub["topic_title"], "photo")
        await message.answer("Принял (без выбора темы). " + praise,
                             reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))
        try:
            # Обновляем данные пользователя перед отправкой уведомления админу
            await ensure_user_with_current_data(message.from_user.id)
            await bot.forward_message(ADMIN_ID, message.chat.id, message.message_id)
        except Exception:
            pass
        return

    await message.answer("Чтобы сдать ДЗ: сначала нажми кнопку в меню, выбери тему, затем отправь фото.",
                         reply_markup=make_main_kb(message.from_user.id == ADMIN_ID))


# media-groups finalizer (runs periodically)
async def process_media_groups():
    now = time.time()
    keys = list(media_groups.keys())
    for key in keys:
        entry = media_groups.get(key)
        if not entry:
            continue
        if now - entry['last_update'] > 1.5:
            try:
                uid = entry.get("uid")
                p_snapshot = entry.get("pending_snapshot") or {}
                p_type = p_snapshot.get("type", "dz")
                section = p_snapshot.get("section", "Без раздела")
                topic = p_snapshot.get("topic", {"id": "none", "title": "Альбом"})
                file_ids = entry.get("file_ids", [])
                caption = entry.get("caption", "") or f"Альбом из {len(file_ids)} фото"
                sub = {"type": p_type, "section": section, "topic_id": topic.get("id"),
                       "topic_title": topic.get("title"),
                       "content_type": "photo_album", "content_summary": caption, "photo_file_id": ";".join(file_ids),
                       "date": today_str(), "ts": datetime.utcnow().isoformat()}
                ensure_user_record_by_id(int(uid))
                fake_user = types.User(id=int(uid), is_bot=False, first_name="", username="")
                add_submission_obj(fake_user, sub, None)
                # save files if conspect
                if sub['type'] == 'conspect':
                    files = []
                    names = []
                    for idx, fid in enumerate(file_ids, start=1):
                        b = await download_file_bytes(fid)
                        if b:
                            name = f"photo_{idx}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.jpg"
                            files.append(b);
                            names.append(name)
                    if files:
                        save_conspect_files(uid, section.replace("/", "_"), sub['topic_id'], files, names)
                try:
                    await bot.send_message(int(uid), get_praise(int(uid), sub['section'], sub['topic_title'], 'photo'))
                except Exception:
                    pass
                try:
                    # Обновляем данные пользователя перед отправкой уведомления админу
                    await ensure_user_with_current_data(int(uid))
                    await bot.send_message(ADMIN_ID,
                                           f"📸 {mention_html_by_id(int(uid))} прислал {sub['type'].upper()} (альбом): {sub['section']} — {sub['topic_title']}",
                                           parse_mode="HTML")
                except Exception:
                    pass
            except Exception:
                logger.exception('Error finalizing media group %s', key)
            finally:
                media_groups.pop(key, None)


# ADMIN HELPERS
async def produce_and_send_user_export(admin_id: int, identifier: str):
    identifier = (identifier or "").lstrip('@').strip()
    target_uid = None
    if identifier.isdigit():
        cursor.execute('SELECT id FROM users WHERE id = ?', (int(identifier),))
        r = cursor.fetchone()
        if r:
            target_uid = str(r[0])
    if not target_uid:
        cursor.execute(
            'SELECT id, username, first_name FROM users WHERE LOWER(username) = LOWER(?) OR LOWER(first_name) = LOWER(?)',
            (identifier, identifier))
        r = cursor.fetchone()
        if r:
            target_uid = str(r[0])

    if not target_uid:
        await bot.send_message(admin_id, "Пользователь не найден.")
        return

    # fetch submissions
    cursor.execute(
        'SELECT type, section, topic_id, topic_title, content_type, content_summary, photo_file_id, date, ts FROM submissions WHERE user_id = ?',
        (int(target_uid),))
    subs = cursor.fetchall()
    if not subs:
        await bot.send_message(admin_id, "У пользователя нет отправлений.")
        return

    cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (int(target_uid),))
    row = cursor.fetchone()
    username = (row[0] or row[1]) if row else target_uid

    lines = [f"Выгрузка для @{username} (id: {target_uid}). Всего: {len(subs)}"]
    for s in subs:
        lines.append(f"- [{s[0].upper()}] {s[7]} {s[1]} — {s[3]} — {s[5]}")
    text = "\n".join(lines)
    for chunk in [text[i:i + 3900] for i in range(0, len(text), 3900)]:
        await bot.send_message(admin_id, chunk)

    rows = []
    for s in subs:
        rows.append({"user_id": target_uid, "username": username, "type": s[0], "section": s[1], "topic_id": s[2],
                     "topic_title": s[3], "content_type": s[4], "content_summary": s[5], "photo_file_id": s[6],
                     "date": s[7], "ts": s[8]})
    excel_bio = io.BytesIO()
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(excel_bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="submissions")
    excel_bio.seek(0)
    await bot.send_document(admin_id, InputFile(excel_bio, filename=f"user_{target_uid}_submissions.xlsx"))

    # prepare zip of saved files
    zip_bio = io.BytesIO()
    with zipfile.ZipFile(zip_bio, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        user_dir = os.path.join(CONSPECTS_DIR, target_uid)
        if os.path.exists(user_dir):
            for root, _, files in os.walk(user_dir):
                for f in files:
                    full = os.path.join(root, f)
                    arc = os.path.relpath(full, user_dir)
                    zf.write(full, arc)
        # download photos referenced in submissions
        for s in subs:
            pf = s[6]
            if pf:
                for fid in str(pf).split(";"):
                    fid = fid.strip()
                    if not fid:
                        continue
                    b = await download_file_bytes(fid)
                    if b:
                        zf.writestr(f"downloaded_{fid[:8]}.jpg", b)
    zip_bio.seek(0)
    if zip_bio.getbuffer().nbytes > 0:
        await bot.send_document(admin_id, InputFile(zip_bio, filename="user_files.zip"))
    else:
        await bot.send_message(admin_id, "У пользователя нет загруженных файлов/фото.")


def cleanup_empty_columns():
    """Очищает пустые столбцы в базе данных"""
    try:
        # Очищаем пустые username
        cursor.execute("UPDATE users SET username = '' WHERE username IS NULL")

        # Очищаем пустые first_name
        cursor.execute("UPDATE users SET first_name = '' WHERE first_name IS NULL")

        # Очищаем пустые значения в submissions
        cursor.execute("UPDATE submissions SET section = '' WHERE section IS NULL")
        cursor.execute("UPDATE submissions SET topic_id = '' WHERE topic_id IS NULL")
        cursor.execute("UPDATE submissions SET topic_title = '' WHERE topic_title IS NULL")
        cursor.execute("UPDATE submissions SET content_type = '' WHERE content_type IS NULL")
        cursor.execute("UPDATE submissions SET content_summary = '' WHERE content_summary IS NULL")
        cursor.execute("UPDATE submissions SET photo_file_id = '' WHERE photo_file_id IS NULL")

        # Очищаем пустые значения в miss_reasons
        cursor.execute("UPDATE miss_reasons SET reason = '' WHERE reason IS NULL")

        conn.commit()
        return True
    except Exception as e:
        logger.exception("Error cleaning up empty columns")
        return False


async def delete_user_submissions(admin_id: int, identifier: str):
    identifier = (identifier or "").lstrip('@').strip()
    target_uid = None

    # Поиск по ID
    if identifier.isdigit():
        cursor.execute('SELECT id FROM users WHERE id = ?', (int(identifier),))
        result = cursor.fetchone()
        if result:
            target_uid = str(result[0])

    # Поиск по username или first_name
    if not target_uid:
        cursor.execute('SELECT id FROM users WHERE LOWER(username) = LOWER(?) OR LOWER(first_name) = LOWER(?)',
                       (identifier, identifier))
        result = cursor.fetchone()
        if result:
            target_uid = str(result[0])

    if not target_uid:
        await bot.send_message(admin_id, "❌ Пользователь не найден.")
        return False

    try:
        # Удаляем все данные пользователя
        cursor.execute('DELETE FROM submissions WHERE user_id = ?', (int(target_uid),))
        cursor.execute('DELETE FROM miss_reasons WHERE user_id = ?', (int(target_uid),))

        # Удаляем самого пользователя из таблицы users
        cursor.execute('DELETE FROM users WHERE id = ?', (int(target_uid),))

        conn.commit()

        # Удаляем файлы пользователя
        user_dir = os.path.join(CONSPECTS_DIR, target_uid)
        if os.path.exists(user_dir):
            try:
                shutil.rmtree(user_dir)
            except Exception as e:
                logger.error(f"Error removing user directory: {e}")

        await bot.send_message(admin_id, f"✅ Данные пользователя {target_uid} полностью удалены.")
        return True

    except Exception as e:
        logger.exception("delete_user_submissions DB error")
        await bot.send_message(admin_id, f"❌ Ошибка при удалении данных пользователя: {str(e)}")
        return False


# ADMIN HANDLERS (callbacks & pending)
@dp.callback_query_handler(lambda c: c.data and c.data.startswith("admin|"))
async def cb_admin(call: types.CallbackQuery):
    if call.from_user.id != ADMIN_ID:
        return await call.answer("Нет доступа", show_alert=True)

    action = call.data.split("|", 1)[1]
    aid = str(call.from_user.id)

    if action == "daily_full":
        await send_daily_excel_to_admin(date.today())
        await call.answer("Отправил дневной отчёт (подробный).")
        return

    if action == "send_daily_now":
        await send_daily_excel_to_admin(date.today())
        await call.answer("Отправил таблицу прямо сейчас.")
        return

    if action == "new_report":
        bio = make_daily_excel(date.today())
        fname = f"daily_report_{date.today().isoformat()}_{int(time.time())}.xlsx"
        await bot.send_document(call.from_user.id, InputFile(bio, filename=fname))
        await call.answer("Создан и отправлен отчёт (сохранён в reports).")
        return

    if action == "cleanup_columns":
        if cleanup_empty_columns():
            await bot.send_message(call.from_user.id, "✅ Пустые столбцы успешно очищены.")
        else:
            await bot.send_message(call.from_user.id, "❌ Ошибка при очистке пустых столбцов.")
        await call.answer()
        return

    if action == "full_history_manual":
        # run update and send PNGs + excel immediately
        cursor.execute('SELECT DISTINCT date FROM submissions ORDER BY date')
        dates = [parse_date(r[0]) for r in cursor.fetchall() if parse_date(r[0])]
        if not dates:
            await bot.send_message(call.from_user.id, "Нет данных для исторического отчёта.")
            await call.answer()
            return
        for d in dates:
            update_full_history_excel(d)
        # create charts inside excel and PNGs & send
        await update_full_history_daily()
        await call.answer("Исторический отчёт и графики отправлены.")
        return

    if action == "delete_reports":
        # delete all files in REPORTS_DIR
        try:
            files = os.listdir(REPORTS_DIR)
            deleted = 0
            for f in files:
                p = os.path.join(REPORTS_DIR, f)
                try:
                    os.remove(p)
                    deleted += 1
                except Exception:
                    pass
            await bot.send_message(call.from_user.id, f"Удалено {deleted} файлов из папки reports.")
        except Exception:
            await bot.send_message(call.from_user.id, "Ошибка при удалении отчётов.")
        await call.answer()
        return

    if action == "export_user":
        admin_pending[aid] = {"action": "export_user"}
        await bot.send_message(call.from_user.id, "Введи ID или username ученика для выгрузки (можно с @):")
        await call.answer()
        return

    if action == "delete_user":
        admin_pending[aid] = {"action": "delete_user"}
        await bot.send_message(call.from_user.id, "Введи ID или username ученика для удаления (можно с @):")
        await call.answer()
        return

    if action == "reset_all":
        try:
            cursor.execute('DELETE FROM submissions')
            cursor.execute('DELETE FROM miss_reasons')
            cursor.execute('DELETE FROM users')
            conn.commit()
            if os.path.exists(CONSPECTS_DIR):
                shutil.rmtree(CONSPECTS_DIR)
                os.makedirs(CONSPECTS_DIR, exist_ok=True)
            await bot.send_message(call.from_user.id, "Все данные и файлы сброшены.")
        except Exception:
            logger.exception("reset_all error")
            await bot.send_message(call.from_user.id, "Ошибка при сбросе данных.")
        await call.answer()
        return

    if action == "cancel":
        await bot.send_message(call.from_user.id, "Операция отменена.", reply_markup=admin_kb())
        await call.answer()
        return


@dp.message_handler(lambda m: str(m.from_user.id) in admin_pending)
async def admin_pending_text(message: types.Message):
    aid = str(message.from_user.id)
    if message.from_user.id != ADMIN_ID:
        return

    task = admin_pending.pop(aid, None)
    if not task:
        return

    action = task.get("action")
    identifier = message.text.strip()

    if action == "export_user":
        await produce_and_send_user_export(message.from_user.id, identifier)
        await message.answer("Выгрузка завершена.", reply_markup=admin_kb())
        return

    if action == "delete_user":
        success = await delete_user_submissions(message.from_user.id, identifier)
        if success:
            await message.answer("✅ Удаление выполнено.", reply_markup=admin_kb())
        else:
            await message.answer("❌ Удаление не выполнено.", reply_markup=admin_kb())
        return


# ADMIN COMMANDS FOR USER MANAGEMENT
@dp.message_handler(commands=['get_user'])
async def cmd_get_user(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return await message.answer("Доступ запрещён")

    try:
        # Получаем ID из команды /get_user 123456789
        args = message.get_args()
        if not args:
            return await message.answer("Использование: /get_user <user_id>")

        user_id = int(args.strip())
        username = get_username_by_id(user_id)
        display_name = get_user_display_name_by_id(user_id)

        response = f"ID: {user_id}\n"
        response += f"Username: @{username}\n" if username else "Username: не установлен\n"
        response += f"Display name: {display_name}\n"
        response += f"Упоминание: {mention_html_by_id(user_id)}"

        await message.answer(response, parse_mode="HTML")

    except ValueError:
        await message.answer("Некорректный ID пользователя")
    except Exception as e:
        logger.exception("Error in cmd_get_user")
        await message.answer("Ошибка при получении информации о пользователе")


@dp.message_handler(commands=['update_user'])
async def cmd_update_user(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return await message.answer("Доступ запрещён")

    try:
        args = message.get_args()
        if not args:
            return await message.answer("Использование: /update_user <user_id>")

        user_id = int(args.strip())
        await ensure_user_with_current_data(user_id)

        username = get_username_by_id(user_id)
        display_name = get_user_display_name_by_id(user_id)

        await message.answer(
            f"Данные пользователя обновлены:\n"
            f"ID: {user_id}\n"
            f"Username: @{username}\n"
            f"Display name: {display_name}",
            parse_mode="HTML"
        )

    except ValueError:
        await message.answer("Некорректный ID пользователя")
    except Exception as e:
        logger.exception("Error in cmd_update_user")
        await message.answer("Ошибка при обновлении данных пользователя")


# ---------------- SCHEDULER TASKS ----------------
async def daily_reminder():
    try:
        cursor.execute('SELECT id FROM users')
        rows = cursor.fetchall()
        for row in rows:
            uid = row[0]
            try:
                await bot.send_message(int(uid),
                                       "⏰ Напоминание: не забудьте сегодня сдать ДЗ и/или конспект. Нажми в меню 'Сдать ДЗ' или 'Сдать конспект'.")
            except Exception:
                pass
    except Exception:
        logger.exception("daily_reminder error")


async def ask_missed_reason():
    """Исправленная функция: спрашивает причину пропуска только у тех, кто ничего не сдал за день"""
    dstr = today_str()
    try:
        # Получаем всех пользователей
        cursor.execute('SELECT id FROM users')
        all_uids = {r[0] for r in cursor.fetchall()}

        # Получаем пользователей, которые сдали что-либо сегодня
        cursor.execute('SELECT DISTINCT user_id FROM submissions WHERE date = ?', (dstr,))
        submitted_uids = {r[0] for r in cursor.fetchall()}

        # Спрашиваем только тех, кто ничего не сдал
        to_ask = all_uids - submitted_uids

        for uid in to_ask:
            try:
                reasons_pending[str(uid)] = dstr
                await bot.send_message(int(uid),
                                       f"Сегодня ({dstr}) ты ничего не сдал(а). Можешь коротко указать причину пропуска? (ответ будет сохранён)")
            except Exception:
                reasons_pending.pop(str(uid), None)
                pass
    except Exception:
        logger.exception("ask_missed_reason error")


async def daily_admin_report():
    d = today_str()
    try:
        cursor.execute('SELECT COUNT(*) FROM submissions WHERE date = ? AND type = ?', (d, 'dz'))
        dz = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM submissions WHERE date = ? AND type = ?', (d, 'conspect'))
        cons = cursor.fetchone()[0]
        text = f"Ежедневный отчёт за {d}:\nДЗ: {dz}\nКонспект: {cons}\nДля подробностей нажми в админ-панели 'Дневной отчёт (подробный)'."
        await bot.send_message(ADMIN_ID, text)
    except Exception:
        logger.exception("daily_admin_report error")


# ---------------- FULL HISTORY, CHARTS AND TOPS ----------------
def update_full_history_excel(target_date: date):
    """
    Append one row per user for target_date into single sheet "ALL".
    Ensures clickable HYPERLINK formulas to tg://user?id=..., fills task_flag,
    skips rows with dz==0 and conspect==0 and empty reason, and avoids duplicates.
    """
    path = os.path.join(REPORTS_DIR, "full_history.xlsx")
    # ensure file exists with sheet ALL
    if not os.path.exists(path):
        # create with header
        from openpyxl import Workbook
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = "ALL"
        headers = ["user_id", "name", "tag", "date", "dz", "conspect", "miss_reason", "task_flag", "MessageLink"]
        for i, h in enumerate(headers, start=1):
            ws_new.cell(row=1, column=i, value=h)
        wb_new.save(path)

    wb = load_workbook(path)
    if "ALL" not in wb.sheetnames:
        wb.create_sheet("ALL")
    ws = wb["ALL"]

    # build set of existing (user_id, date) to avoid duplicates
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            uid = row[0]  # user_id is now first column
            d = row[3]  # date is now 4th column
            existing.add((str(uid), str(d)))
        except Exception:
            pass

    dstr = target_date.isoformat()

    # iterate users from DB and compute statuses
    try:
        cursor.execute('SELECT id, username, first_name FROM users ORDER BY id')
        users = cursor.fetchall()
    except Exception:
        users = []

    rows_to_append = []
    for u in users:
        uid, uname, fname = u
        uid_s = str(uid)
        # skip if already present for that date
        if (uid_s, dstr) in existing:
            continue
        # compute dz and cons presence
        try:
            cursor.execute('SELECT COUNT(*) FROM submissions WHERE user_id = ? AND date = ? AND type = ?',
                           (uid, dstr, 'dz'))
            dz = cursor.fetchone()[0] or 0
            cursor.execute('SELECT COUNT(*) FROM submissions WHERE user_id = ? AND date = ? AND type = ?',
                           (uid, dstr, 'conspect'))
            cons = cursor.fetchone()[0] or 0
            cursor.execute('SELECT reason FROM miss_reasons WHERE user_id = ? AND date = ?', (uid, dstr))
            rr = cursor.fetchone()
            reason = rr[0] if rr and rr[0] else ""
            # task_flag: last submission topic_id or topic_title
            cursor.execute(
                'SELECT topic_id, topic_title FROM submissions WHERE user_id = ? AND date = ? ORDER BY ts DESC LIMIT 1',
                (uid, dstr))
            trow = cursor.fetchone()
            task_flag = ""
            if trow:
                task_flag = trow[0] or trow[1] or ""
        except Exception:
            dz = 0;
            cons = 0;
            reason = "";
            task_flag = ""

        # skip empty rows: dz==0 and cons==0 and no reason
        if (int(dz) == 0) and (int(cons) == 0) and (not str(reason).strip()):
            continue

        # build name and tag
        # name: first_name or user_123 - but now as clickable link
        if fname and str(fname).strip():
            name_value = fname
        else:
            name_value = f"user_{uid}"

        # Create clickable name using HYPERLINK formula
        clickable_name = f'=HYPERLINK("tg://user?id={uid}", "{name_value}")'

        # tag: @username if exists else ""
        if uname and str(uname).strip():
            tag_value = f"@{uname}"
        else:
            tag_value = ""

        # message link
        try:
            cursor.execute('SELECT message_id FROM submissions WHERE user_id = ? AND date = ? ORDER BY ts DESC LIMIT 1',
                           (uid, dstr))
            mid_row = cursor.fetchone()
            mid = mid_row[0] if mid_row and mid_row[0] else None
        except Exception:
            mid = None

        if mid:
            msg_link = f'=HYPERLINK("tg://openmessage?chat_id={uid}&message_id={mid}", "Open")'
        else:
            msg_link = ""

        rows_to_append.append(
            (uid, clickable_name, tag_value, dstr, int(dz), int(cons), reason or "", task_flag or "", msg_link))

    # append rows to sheet
    for r in rows_to_append:
        next_row = ws.max_row + 1
        for ci, val in enumerate(r, start=1):
            cell = ws.cell(row=next_row, column=ci, value=val)
            # For name column (column 2), set as formula if it contains HYPERLINK
            if ci == 2 and isinstance(val, str) and 'HYPERLINK' in val:
                cell.value = val  # This will be interpreted as formula by Excel

    # style worksheet and autosize
    try:
        style_worksheet(ws)
        ws.freeze_panes = ws['A2']
    except Exception:
        pass

    wb.save(path)


def generate_miss_graph_by_student_png() -> _BytesIO:
    """Generate PNG showing number of misses per student (overall) and return BytesIO."""
    # count misses from miss_reasons table and also infer misses by days without submissions
    try:
        # Number of days tracked
        cursor.execute('SELECT COUNT(DISTINCT date) FROM submissions')
        days_row = cursor.fetchone()
        total_days = days_row[0] if days_row and days_row[0] else 0

        # Count misses recorded in miss_reasons
        cursor.execute('SELECT user_id, COUNT(*) as cnt FROM miss_reasons GROUP BY user_id ORDER BY cnt DESC')
        rows = cursor.fetchall()
        users = []
        counts = []
        for uid, cnt in rows:
            cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (uid,))
            r = cursor.fetchone()
            name = (('@' + r[0]) if r and r[0] else (r[1] if r and r[1] else f'user_{uid}'))
            users.append(name)
            counts.append(cnt)

        # If no explicit miss_reasons, try infer by checking days without submissions per user
        if not rows:
            cursor.execute('SELECT id, username, first_name FROM users')
            all_users = cursor.fetchall()
            for u in all_users:
                uid, uname, fname = u
                cursor.execute('SELECT COUNT(DISTINCT date) FROM submissions WHERE user_id = ?', (uid,))
                sub_days = cursor.fetchone()[0] or 0
                misses = max(0, total_days - sub_days) if total_days > 0 else 0
                users.append(('@' + uname) if uname else (fname or f'user_{uid}'))
                counts.append(misses)

        # plot bar chart
        fig, ax = plt.subplots(figsize=(8, max(4, len(users) * 0.4)))
        ax.bar(range(len(users)), counts)
        ax.set_xticks(range(len(users)))
        ax.set_xticklabels(users, rotation=45, ha='right')
        ax.set_ylabel('Пропуски (кол-во дней)')
        ax.set_title('Пропуски по ученикам (всего)')
        plt.tight_layout()

        bio = _BytesIO()
        fig.savefig(bio, format='png', bbox_inches='tight')
        plt.close(fig)
        bio.seek(0)
        return bio
    except Exception:
        logger.exception("generate_miss_graph_by_student_png error")
        return _BytesIO()


def generate_top_students_png(kind: str = "dz", top_n: int = 10) -> _BytesIO:
    """Generate PNG of top students by submissions of `kind` ('dz' or 'conspect')."""
    try:
        cursor.execute(
            'SELECT user_id, COUNT(*) as cnt FROM submissions WHERE type = ? GROUP BY user_id ORDER BY cnt DESC LIMIT ?',
            (kind, top_n))
        rows = cursor.fetchall()
        users = []
        counts = []
        for uid, cnt in rows:
            cursor.execute('SELECT username, first_name FROM users WHERE id = ?', (uid,))
            r = cursor.fetchone()
            name = (('@' + r[0]) if r and r[0] else (r[1] if r and r[1] else f'user_{uid}'))
            users.append(name)
            counts.append(cnt)

        if not rows:
            bio = _BytesIO()
            fig, ax = plt.subplots(figsize=(6, 3))
            ax.text(0.5, 0.5, 'Нет данных', ha='center', va='center')
            plt.axis('off')
            fig.savefig(bio, format='png', bbox_inches='tight')
            plt.close(fig)
            bio.seek(0)
            return bio

        fig, ax = plt.subplots(figsize=(8, max(3, len(users) * 0.4)))
        ax.bar(range(len(users)), counts)
        ax.set_xticks(range(len(users)))
        ax.set_xticklabels(users, rotation=45, ha='right')
        ax.set_ylabel('Кол-во отправлений')
        ax.set_title('Топ учеников по ' + ('ДЗ' if kind == 'dz' else 'конспектам'))
        plt.tight_layout()

        bio = _BytesIO()
        fig.savefig(bio, format='png', bbox_inches='tight')
        plt.close(fig)
        bio.seek(0)
        return bio
    except Exception:
        logger.exception("generate_top_students_png error")
        return _BytesIO()


async def update_full_history_daily():
    """Regenerate full_history.xlsx (sheets for each date), charts and PNGs, and send to ADMIN."""
    try:
        cursor.execute('SELECT DISTINCT date FROM submissions ORDER BY date')
        dates = [parse_date(r[0]) for r in cursor.fetchall() if parse_date(r[0])]
        if not dates:
            return

        for d in dates:
            update_full_history_excel(d)

        # create charts inside excel: summary sheets for DZ & conspects
        path = os.path.join(REPORTS_DIR, "full_history.xlsx")
        wb = load_workbook(path)

        # remove old summary sheets if present
        for name in ("Chart_DZ", "Chart_Conspects"):
            if name in wb.sheetnames:
                wb.remove(wb[name])

        # DZ summary
        cursor.execute("SELECT date, COUNT(*) FROM submissions WHERE type='dz' GROUP BY date ORDER BY date")
        dz_rows = cursor.fetchall()
        ws1 = wb.create_sheet("Chart_DZ")
        ws1.append(["date", "dz_count"])
        for r in dz_rows:
            ws1.append([r[0], r[1]])

        # Conspects summary
        cursor.execute("SELECT date, COUNT(*) FROM submissions WHERE type='conspect' GROUP BY date ORDER BY date")
        c_rows = cursor.fetchall()
        ws2 = wb.create_sheet("Chart_Conspects")
        ws2.append(["date", "cons_count"])
        for r in c_rows:
            ws2.append([r[0], r[1]])

        wb.save(path)

        # generate PNGs
        miss_png = generate_miss_graph_by_student_png()
        top_dz_png = generate_top_students_png('dz')
        top_cons_png = generate_top_students_png('conspect')

        # send to admin: three PNGs as documents
        try:
            await bot.send_photo(ADMIN_ID, photo=InputFile(miss_png, filename='misses_by_student.png'))
        except Exception:
            try:
                await bot.send_document(ADMIN_ID, InputFile(miss_png, filename='misses_by_student.png'))
            except Exception:
                pass
        try:
            await bot.send_photo(ADMIN_ID, photo=InputFile(top_dz_png, filename='top_dz.png'))
        except Exception:
            try:
                await bot.send_document(ADMIN_ID, InputFile(top_dz_png, filename='top_dz.png'))
            except Exception:
                pass
        try:
            await bot.send_photo(ADMIN_ID, photo=InputFile(top_cons_png, filename='top_cons.png'))
        except Exception:
            try:
                await bot.send_document(ADMIN_ID, InputFile(top_cons_png, filename='top_cons.png'))
            except Exception:
                pass

        # finally send the excel file
        try:
            await bot.send_document(ADMIN_ID, InputFile(path, filename='full_history.xlsx'))
        except Exception:
            pass

    except Exception:
        logger.exception("update_full_history_daily error")


async def on_startup(dispatcher):
    scheduler.add_job(daily_reminder, "cron", hour=18, minute=0)
    scheduler.add_job(daily_admin_report, "cron", hour=23, minute=55)
    scheduler.add_job(process_media_groups, "interval", seconds=2)
    scheduler.add_job(ask_missed_reason, "cron", hour=23, minute=57)
    # schedule full history daily update at 23:50
    scheduler.add_job(update_full_history_daily, 'cron', hour=23, minute=50)
    scheduler.start()
    logger.info("Scheduler started")


# - MAIN
if __name__ == "__main__":
    print("Starting bot. Make sure API_TOKEN and ADMIN_ID are set in env.")
    print(f"ADMIN_ID: {ADMIN_ID}")
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)


